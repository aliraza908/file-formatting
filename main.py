import pandas as pd
import openpyxl

file_path = r"C:\Users\taree\OneDrive\Desktop\Project\data_report.xlsx"
fp = r"C:\Users\taree\OneDrive\Desktop\Project\dummy.xlsx"
sheet_name_to_delete = 'PS Amazon'
sheet_name_to_shift = 'Data'
workbook =openpyxl.load_workbook(file_path)

if sheet_name_to_delete and sheet_name_to_shift in workbook.sheetnames:
    sheet_to_delete = workbook[sheet_name_to_delete]
    sheet_to_shift = workbook[sheet_name_to_shift]
    workbook.remove(sheet_to_delete)
    workbook.remove(sheet_to_shift)
    workbook._add_sheet(sheet_to_shift)
    workbook.save(fp)
#    print(f"The sheet {sheet_name_to_delete} has been deleted.")
#    print(f"The sheet {sheet_name_to_shift} has been shifted to the end.")



else:
    print("The Sheet doesnot exist in excel file")

df = pd.read_excel(r"C:\Users\taree\OneDrive\Desktop\Project\dummy.xlsx")

column_name_mapping = {
    'Sum': 'Total Item',
    'Matched': 'Price Match',
    'High': 'High Price',
    'Low': 'Low Price',
}

df.rename(columns=column_name_mapping, inplace=True)

column1_name = 'Listed'
column2_name = 'Not Listed'

col1_temp = df[column1_name].copy()
df[column1_name] = df[column2_name] #assigned second column to first column
df[column2_name] = col1_temp

df.rename(columns={column1_name: column2_name, column2_name: column1_name}, inplace=True)

#In this code, we use .str.contains('|'.join(key), case=False) to check if the 'Channel' column contains any of the words in the key list (using the '|' as a separator between the words in the regex pattern). The case=False argument is used to perform a case-insensitive match. The result is a boolean Series that identifies the rows to be removed from the DataFrame. We then use the .drop() method to remove those rows from the DataFrame df.
key = ['psa listings', 'unsold']
rows_to_remove = df[df['Channel'].str.contains('|'.join(key), case=False)]
df = df.drop(rows_to_remove.index)

values_to_rename = {
    'bsa listings': 'BSAmazon Listings',
    'psw item report': 'PSWalmart Item Report',
    'bsw item report': 'BSWalmart Item Report',
    'pse active': 'PSeBay Active',
    'baabs active': 'Baabs Active',
    'mmm active': 'Mmm Active',
    'bse active': 'BSeBay Active'

}

df['Channel'] = df['Channel'].replace(values_to_rename)


# Define the original Excel file path and the sheet name you want to modify
original_excel_path = r"C:\Users\taree\OneDrive\Desktop\Project\dummy.xlsx"
sheet_name_to_modify = "Report"

# Read the original Excel file
book = openpyxl.load_workbook(original_excel_path)

# Get the sheet you want to modify
sheet_to_modify = book[sheet_name_to_modify]

# Convert the modified DataFrame df to a Pandas DataFrame
df_modified = pd.DataFrame(df)

# Clear the existing contents in the sheet except for the first row (header row)
for row in sheet_to_modify.iter_rows(min_row=2, min_col=1):
    for cell in row:
        cell.value = None

# Write the column names to the first row in the sheet
for col_index, column_name in enumerate(df_modified.columns, 1):
    sheet_to_modify.cell(row=1, column=col_index, value=column_name)

# Write the row values to the sheet starting from cell A2
for row_index, row in df_modified.iterrows():
    for col_index, value in enumerate(row, 1):
        sheet_to_modify.cell(row=row_index + 1, column=col_index, value=value)

# Save the modified Excel file with the changes to the "Report" sheet
book.save(original_excel_path)

